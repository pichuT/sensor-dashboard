"""
convert_to_excel.py — CubeSat SOAR26
Converts one or more sensor_*.txt files from the SD card into
balloon_flight_data.xlsx for the React sensor dashboard.
 
USAGE:
  python convert_to_excel.py sensor_41817.txt
  python convert_to_excel.py sensor_*.txt        (multiple files = multiple rows)
  python convert_to_excel.py                     (auto-finds all sensor_*.txt in current folder)
 
OUTPUT:
  balloon_flight_data.xlsx  (drop this into your dashboard's public/ folder)
"""
 
import sys
import re
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
 
def parse_sensor_file(filepath):
    """Parse a sensor_*.txt file and return a dict of values."""
    data = {
        "Timestamp":    None,
        "Temperature":  None,
        "Pressure":     None,
        "Humidity":     None,
        "Altitude":     None,
        "Accel_X":      None,
        "Accel_Y":      None,
        "Accel_Z":      None,
        "Latitude":     None,
        "Longitude":    None,
        "GPS_Altitude": None,
        "Satellites":   None,
        "GPS_Valid":    None,
    }
 
    # Use the file's modification time as a timestamp if no time in file
    mtime = os.path.getmtime(filepath)
    data["Timestamp"] = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
 
    with open(filepath, "r") as f:
        content = f.read()
 
    # Parse each field using regex
    def extract(pattern, text, group=1, cast=float):
        m = re.search(pattern, text)
        if m:
            try:
                return cast(m.group(group))
            except:
                return m.group(group)
        return None
 
    data["Temperature"]  = extract(r"Temperature:([\d.\-]+)", content)
    data["Pressure"]     = extract(r"Pressure:([\d.\-]+)", content)
    data["Humidity"]     = extract(r"Humidity:([\d.\-]+)", content)
    data["Altitude"]     = extract(r"Altitude:([\d.\-]+)", content)
    data["Accel_X"]      = extract(r"X:([\d.\-]+)", content)
    data["Accel_Y"]      = extract(r"Y:([\d.\-]+)", content)
    data["Accel_Z"]      = extract(r"Z:([\d.\-]+)", content)
    data["Latitude"]     = extract(r"Latitude:([\d.\-]+)", content)
    data["Longitude"]    = extract(r"Longitude:([\d.\-]+)", content)
    data["GPS_Altitude"] = extract(r"GPS_Altitude:([\d.\-]+)", content)
    data["Satellites"]   = extract(r"Satellites:(\d+)", content, cast=int)
    data["GPS_Valid"]    = extract(r"GPS_Valid:(\w+)", content, cast=str)
 
    return data
 
def convert(filepaths, output="balloon_flight_data.xlsx"):
    rows = [parse_sensor_file(fp) for fp in filepaths]
 
    headers = [
        "Timestamp", "Temperature", "Pressure", "Humidity",
        "Altitude", "Accel_X", "Accel_Y", "Accel_Z",
        "Latitude", "Longitude", "GPS_Altitude", "Satellites", "GPS_Valid"
    ]
 
    wb = Workbook()
    ws = wb.active
    ws.title = "SensorData"
 
    # Header row styling
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
 
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18
 
    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col, value=row.get(header))
 
    wb.save(output)
    print(f"Saved {len(rows)} row(s) to {output}")
    print(f"Drop it into your dashboard's public/ folder and refresh!")
 
if __name__ == "__main__":
    # Get file paths from args or auto-discover
    if len(sys.argv) > 1:
        filepaths = sys.argv[1:]
    else:
        filepaths = sorted(glob.glob("sensor_*.txt"))
        if not filepaths:
            print("No sensor_*.txt files found. Pass filenames as arguments.")
            sys.exit(1)
 
    print(f"Found {len(filepaths)} file(s): {', '.join(filepaths)}")
    convert(filepaths)